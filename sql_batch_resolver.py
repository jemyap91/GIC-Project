"""Batch job dependency resolver using SQL (SQLite).

Resolves execution order for batch jobs by analyzing dependency rules
and computing execution levels using recursive CTEs. Steps at the same
level can run in parallel.
"""
import sqlite3
from typing import Optional


class BatchResolver:
    """Resolves batch job step execution order from dependency rules."""

    def __init__(self, db_path: str = ":memory:"):
        self.conn = sqlite3.connect(db_path)

    def create_tables(self):
        self.conn.executescript("""
            CREATE TABLE IF NOT EXISTS prog_name (
                unit_nbr INTEGER NOT NULL,
                step_seq_id INTEGER NOT NULL,
                step_prog_name TEXT NOT NULL,
                PRIMARY KEY (unit_nbr, step_seq_id)
            );

            CREATE TABLE IF NOT EXISTS dependency_rules (
                unit_nbr INTEGER NOT NULL,
                rule_id INTEGER NOT NULL,
                step_seq_id INTEGER NOT NULL,
                step_dep_id INTEGER NOT NULL,
                PRIMARY KEY (unit_nbr, rule_id)
            );
        """)

    def load_prog_names(self, data: list[tuple]):
        self.conn.executemany(
            "INSERT INTO prog_name (unit_nbr, step_seq_id, step_prog_name) VALUES (?, ?, ?)",
            data,
        )
        self.conn.commit()

    def load_dependency_rules(self, data: list[tuple]):
        self.conn.executemany(
            "INSERT INTO dependency_rules (unit_nbr, rule_id, step_seq_id, step_dep_id) VALUES (?, ?, ?, ?)",
            data,
        )
        self.conn.commit()

    def load_from_excel(self, filepath: str):
        import openpyxl

        wb = openpyxl.load_workbook(filepath)

        ws = wb["PROG_NAME"]
        prog_data = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] is not None:
                prog_data.append((row[0], row[1], row[2]))
        self.load_prog_names(prog_data)

        ws = wb["DEPENDENCY_RULES"]
        dep_data = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] is not None:
                dep_data.append((row[0], row[1], row[2], row[3]))
        self.load_dependency_rules(dep_data)

    def get_sql_query(self) -> str:
        """Return the raw SQL query used for dependency resolution.

        This is the core SQL solution using a recursive CTE to compute
        execution levels from dependency rules.
        """
        return """
WITH RECURSIVE step_levels AS (
    -- Base case: steps with no real dependency (dep_id = 0)
    SELECT
        d.unit_nbr,
        d.step_seq_id,
        0 AS exec_level
    FROM dependency_rules d
    WHERE d.unit_nbr = :unit_nbr
      AND d.step_dep_id = 0

    UNION ALL

    -- Recursive case: step level = parent level + 1
    SELECT
        d.unit_nbr,
        d.step_seq_id,
        sl.exec_level + 1
    FROM dependency_rules d
    JOIN step_levels sl
      ON d.step_dep_id = sl.step_seq_id
     AND d.unit_nbr = sl.unit_nbr
    WHERE d.step_dep_id != 0
),
-- A step's actual level is the MAX of all paths to it
-- This ensures a step waits for ALL dependencies, not just the first
max_levels AS (
    SELECT
        unit_nbr,
        step_seq_id,
        MAX(exec_level) AS exec_level
    FROM step_levels
    GROUP BY unit_nbr, step_seq_id
)
SELECT
    ml.unit_nbr,
    ml.exec_level,
    ml.step_seq_id,
    p.step_prog_name
FROM max_levels ml
JOIN prog_name p
  ON ml.unit_nbr = p.unit_nbr
 AND ml.step_seq_id = p.step_seq_id
ORDER BY ml.exec_level, ml.step_seq_id;
"""

    def resolve_order(self, unit_nbr: int) -> list[tuple]:
        """Resolve execution order using SQL recursive CTE.

        Returns list of (unit_nbr, level, step_seq_id, step_prog_name)
        sorted by level then step_seq_id.
        """
        query = """
            WITH RECURSIVE step_levels AS (
                -- Base case: steps with no real dependency (dep_id = 0)
                SELECT
                    d.unit_nbr,
                    d.step_seq_id,
                    0 AS exec_level
                FROM dependency_rules d
                WHERE d.unit_nbr = ?
                  AND d.step_dep_id = 0

                UNION ALL

                -- Recursive case: step level = parent level + 1
                SELECT
                    d.unit_nbr,
                    d.step_seq_id,
                    sl.exec_level + 1
                FROM dependency_rules d
                JOIN step_levels sl
                  ON d.step_dep_id = sl.step_seq_id
                 AND d.unit_nbr = sl.unit_nbr
                WHERE d.step_dep_id != 0
            ),
            -- A step's actual level is the MAX of all paths to it
            max_levels AS (
                SELECT
                    unit_nbr,
                    step_seq_id,
                    MAX(exec_level) AS exec_level
                FROM step_levels
                GROUP BY unit_nbr, step_seq_id
            )
            SELECT
                ml.unit_nbr,
                ml.exec_level,
                ml.step_seq_id,
                p.step_prog_name
            FROM max_levels ml
            JOIN prog_name p
              ON ml.unit_nbr = p.unit_nbr
             AND ml.step_seq_id = p.step_seq_id
            ORDER BY ml.exec_level, ml.step_seq_id
        """
        cursor = self.conn.execute(query, (unit_nbr,))
        return cursor.fetchall()

    def format_execution_plan(self, unit_nbr: int) -> str:
        """Format execution order as a readable string with level grouping.

        Output shows UNIT_NBR and STEP_PROG_NAME in correct sequence,
        grouped by execution level. Steps within the same level can run
        in parallel.
        """
        results = self.resolve_order(unit_nbr)
        if not results:
            return "No steps found."

        lines = []
        lines.append(f"{'UNIT_NBR':<12}{'STEP_SEQ_ID':<14}{'STEP_PROG_NAME':<55}{'EXEC_LEVEL'}")
        lines.append("-" * 95)

        current_level = -1
        for unit, level, seq_id, prog_name in results:
            if level != current_level:
                if current_level != -1:
                    lines.append("")  # blank line between levels
                current_level = level
            lines.append(f"{unit:<12}{seq_id:<14}{prog_name:<55}{level}")

        return "\n".join(lines)


if __name__ == "__main__":
    import sys

    filepath = sys.argv[1] if len(sys.argv) > 1 else "SQL_Test_1.xlsx"
    resolver = BatchResolver()
    resolver.create_tables()
    resolver.load_from_excel(filepath)

    units = resolver.conn.execute(
        "SELECT DISTINCT unit_nbr FROM prog_name ORDER BY unit_nbr"
    ).fetchall()

    for (unit_nbr,) in units:
        print(resolver.format_execution_plan(unit_nbr))
        print()

    print("=== SQL Query Used ===")
    print(resolver.get_sql_query())
